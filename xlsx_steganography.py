from openpyxl import load_workbook
from openpyxl.styles import Font
from steganograpy_utility import to_bin


def encode(path, msg, bit):
    wb = load_workbook(path)
    sheet = wb.active
    msg += "====="
    binary_msg = to_bin(msg)
    index = 0
    length = len(binary_msg)
    max_length = (sheet.max_row-1) * sheet.max_column * bit
    if length > max_length:
        raise ValueError(f"[!] Message length ({length}) exceeded {max_length}, please adjust the message or bit.")
    for row in sheet.iter_rows():
        if index >= length:
            break
        for cell in row:
            if index < length:
                # get current color and fill type
                try:
                    binary_fill = bin(int(cell.font.color.rgb, 16))[2:].zfill(8)
                except TypeError:
                    binary_fill = '00000000'
                data = ""
                for b in range(bit):
                    if index < length:
                        data += binary_msg[index]
                        index += 1
                    else:
                        data += binary_fill[-bit + b]
                # change the binary_fill
                binary_fill = binary_fill[:-bit] + data
                cell.font = Font(color=hex(int(binary_fill, 2))[2:].zfill(6))
            else:
                break
    return wb


def decode(path, bit):
    wb = load_workbook(path)
    sheet = wb.active
    binary_msg = ""
    msg = ""
    for row in sheet.iter_rows():
        for cell in row:
            try:
                color = cell.font.color.rgb
            except:
                raise ValueError("No Steganography found in Excel file")
            binary_color = bin(int(color, 16))[2:].zfill(8)
            binary_msg += binary_color[-bit:]
            if len(binary_msg) >= 8:
                msg += chr(int(binary_msg[:8], 2))
                if msg.endswith("====="):
                    return msg[:-5]
                binary_msg = binary_msg[8:]


def save(path, wb):
    if wb:
        wb.save(path)


# save('testingfiles/out.xlsx', encode('testingfiles/SampleData.xlsx', """
# hello world
# hello world
# hello world
# hello world
# hello world
# hello world
# hello world
# hello world
# hello world
# hello world
# hello world
# hello world
# """, 4))
# print(decode('testingfiles/out.xlsx', 4))
